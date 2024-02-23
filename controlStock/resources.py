from django.shortcuts import render, redirect
from .models import *
from openpyxl import load_workbook
from django.contrib import messages

# DB
from django.db.utils import IntegrityError
from django.db.models import Case, When
"""
from django.shortcuts import render, redirect
from .models import *
from .utils import *
from django.contrib import messages
from datetime import datetime
import pytz
# DB
from django.db.utils import IntegrityError
from django.db.models import Case, When
# Login
from django.contrib.auth.decorators import login_required
# Create your views here.
from accounts.views import user_has_required_group, access_denied
from django.contrib.auth.decorators import user_passes_test
"""

# Create your views here.

def inicio(request):
    return render(request, 'inicio.html')

# PRODUCTO
def producto(request):
    producto = Producto.objects.all()
    return render(request, 'producto\\producto.html', {'producto': producto})

def registrarProducto(request):
    if request.method == 'GET':
        return render(request, 'producto\\producto.html')
    else:
        cod = request.POST['cod']
        des = request.POST['des']
        rub = request.POST['rub']
        pre = request.POST['pre']
        try:
            producto = Producto.objects.create(cod=cod, des = des, rub = rub, pre = pre)
            messages.success(request, '¡Registrado!')
        except IntegrityError as e:
            messages.error(request, '¡Ya existe un Producto con ese Codigo!')
        except Exception as e:
            messages.error(request, f"¡Ocurrió un error! {e}")
    return redirect('producto')

def eliminarProducto(request, id_pd):
    producto = Producto.objects.get(id_pd=id_pd)
    producto.delete()
    messages.success(request, '¡Eliminado!')
    return redirect('producto')

def edicionProducto(request, id_pd):
    producto = Producto.objects.get(id_pd=id_pd)
    return render(request, "producto\\edicionProducto.html", {"producto": producto})

def editarProducto(request):
    id_pd = request.POST['id_pd']
    cod = request.POST['cod']
    des = request.POST['des']
    rub = request.POST['rub']
    pre = request.POST['pre']
    #-
    producto = Producto.objects.get(id_pd=id_pd)
    producto.cod = cod
    producto.des = des
    producto.rub = rub
    producto.pre = pre
    producto.save()
    messages.success(request, '¡Datos actualizados!')
    return redirect('producto')



def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            cod, des, pre = row
            Producto.objects.create(cod=cod, des=des, pre=pre)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')
